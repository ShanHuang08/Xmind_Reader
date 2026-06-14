"""Human-friendly Excel exporter for extracted knowledge."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "ID",
    "Module",
    "Path",
    "Scenario",
    "Tags",
    "Precondition",
    "Steps",
    "Expected Result",
    "DB Check",
    "Source XMind File",
    "Source Sheet",
]


def export_excel(cases: list[dict[str, Any]], output_path: Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Knowledge"
    worksheet.append(HEADERS)
    _style_header(worksheet)

    for case in cases:
        source = case.get("source", {})
        worksheet.append(
            [
                case.get("id", ""),
                case.get("module", ""),
                case.get("path", ""),
                case.get("scenario", ""),
                ", ".join(case.get("tags", [])),
                case.get("preconditions", ""),
                _numbered(case.get("steps", [])),
                _numbered(case.get("expected_results", [])),
                _numbered(case.get("db_checks", [])),
                source.get("xmind_file", ""),
                source.get("sheet", ""),
            ]
        )

    _format_sheet(worksheet)
    workbook.save(output_path)
    return output_path


def _numbered(values: list[str]) -> str:
    return "\n".join(f"[{index}]{value}" for index, value in enumerate(values, start=1))


def _style_header(worksheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _format_sheet(worksheet) -> None:
    widths = [16, 18, 42, 38, 28, 46, 54, 54, 44, 28, 24]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
